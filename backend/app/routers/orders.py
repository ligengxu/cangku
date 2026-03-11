from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from app.database import get_db
from app.models import FruitPurchase, SimpleMaterialPurchase, CartonBoxPurchase, CartonBox, Fruit, Supplier
from app.models.user import User
from app.schemas.orders import (
    FruitPurchaseCreate, FruitPurchaseUpdate, FruitPurchaseOut,
    MaterialPurchaseCreate, MaterialPurchaseUpdate, MaterialPurchaseOut,
    CartonPurchaseCreate, CartonPurchaseUpdate, CartonPurchaseOut,
    PaymentStatusUpdate,
)
from app.schemas.common import ApiResponse, PaginatedResponse
from app.middleware.auth import get_current_user, require_admin
from app.utils.cache import cache_clear_prefix
from app.utils.log_action import log_action
from app.services.finance_bridge import push_payment_sync
import csv
import io
import logging

finance_logger = logging.getLogger("finance_bridge")

router = APIRouter(prefix="/orders", tags=["订单管理"])


# ─── 水果采购 ───
@router.get("/fruit", response_model=PaginatedResponse[FruitPurchaseOut])
def list_fruit_purchases(
    page: int = 1,
    page_size: int = 20,
    fruit_name: str | None = None,
    supplier_name: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    payment_status: str | None = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(FruitPurchase).filter(FruitPurchase.deleted_at.is_(None))
    if fruit_name:
        q = q.filter(FruitPurchase.fruit_name.like(f"%{fruit_name}%"))
    if supplier_name:
        q = q.filter(FruitPurchase.supplier_name.like(f"%{supplier_name}%"))
    if start_date:
        q = q.filter(FruitPurchase.purchase_date >= start_date)
    if end_date:
        q = q.filter(FruitPurchase.purchase_date <= end_date)
    if payment_status:
        q = q.filter(FruitPurchase.payment_status == payment_status)

    total = q.count()
    items = q.order_by(desc(FruitPurchase.id)).offset((page - 1) * page_size).limit(page_size).all()
    return PaginatedResponse(data=items, total=total, page=page, page_size=page_size)


@router.post("/fruit", response_model=ApiResponse[FruitPurchaseOut])
def create_fruit_purchase(
    req: FruitPurchaseCreate,
    force: bool = False,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if req.purchase_price <= 0:
        raise HTTPException(status_code=400, detail="采购单价必须大于0")
    if req.purchase_weight <= 0:
        raise HTTPException(status_code=400, detail="采购重量必须大于0")
    if req.purchase_price > 500:
        raise HTTPException(status_code=400, detail="采购单价异常（超过500元/kg），请检查输入")
    if req.purchase_weight > 50000:
        raise HTTPException(status_code=400, detail="采购重量异常（超过50吨），请检查输入")
    fruit = db.query(Fruit).filter(Fruit.id == req.fruit_id).first() if req.fruit_id else None
    supplier = db.query(Supplier).filter(Supplier.id == req.supplier_id, Supplier.deleted_at.is_(None)).first() if req.supplier_id else None

    if req.fruit_id and not fruit:
        raise HTTPException(status_code=400, detail="水果品种不存在")
    if req.supplier_id and not supplier:
        raise HTTPException(status_code=400, detail="供应商不存在")
    if not req.fruit_name and not fruit:
        raise HTTPException(status_code=400, detail="请选择水果或填写水果名称")
    if not req.supplier_name and not supplier:
        raise HTTPException(status_code=400, detail="请选择供应商或填写供应商名称")

    if fruit and not req.fruit_name:
        req.fruit_name = fruit.name
    if supplier and not req.supplier_name:
        req.supplier_name = supplier.name
    if fruit and not req.fruit_id:
        req.fruit_id = fruit.id
    if supplier and not req.supplier_id:
        req.supplier_id = supplier.id

    if not force:
        dup = db.query(FruitPurchase).filter(
            FruitPurchase.supplier_id == req.supplier_id,
            FruitPurchase.fruit_id == req.fruit_id,
            FruitPurchase.purchase_date == req.purchase_date,
            FruitPurchase.purchase_weight == req.purchase_weight,
            FruitPurchase.purchase_price == req.purchase_price,
            FruitPurchase.deleted_at.is_(None),
        ).first()
        if dup:
            raise HTTPException(status_code=409, detail=f"检测到重复采购记录（ID #{dup.id}，同日同供应商同水果同价格同重量），如需继续请确认")

    warnings = []
    recent_avg = db.query(func.avg(FruitPurchase.purchase_price)).filter(
        FruitPurchase.fruit_id == req.fruit_id,
        FruitPurchase.deleted_at.is_(None),
    ).scalar()
    if recent_avg:
        avg_price = float(recent_avg)
        diff_pct = abs(float(req.purchase_price) - avg_price) / avg_price * 100 if avg_price > 0 else 0
        if diff_pct > 30:
            direction = "高于" if float(req.purchase_price) > avg_price else "低于"
            warnings.append(f"本次采购价 ¥{req.purchase_price}/kg {direction}历史均价 ¥{avg_price:.2f}/kg {diff_pct:.0f}%")

    purchase = FruitPurchase(**req.model_dump())
    db.add(purchase)
    log_action(db, user, f"新建水果采购：{req.fruit_name or ''} {req.purchase_weight}kg ¥{req.purchase_price}/kg")
    db.commit()
    db.refresh(purchase)
    cache_clear_prefix("dashboard")

    total_amount = float(req.purchase_price * req.purchase_weight)
    finance_ok = False
    try:
        result = push_payment_sync(
            supplier_name=req.supplier_name or "",
            amount=total_amount,
            reason=f"水果采购：{req.fruit_name or ''} {req.purchase_weight}kg ¥{req.purchase_price}/kg",
            source_order_id=purchase.id,
            source_order_type="fruit",
        )
        finance_ok = result.get("success", False)
    except Exception as e:
        finance_logger.error(f"Push fruit payment to finance failed: {e}")

    msg = "创建成功"
    if not finance_ok:
        msg += "（财务系统推送失败，请联系管理员）"
    if warnings:
        msg += "（" + "；".join(warnings) + "）"
    return ApiResponse(data=purchase, message=msg)


@router.put("/fruit/{purchase_id}", response_model=ApiResponse[FruitPurchaseOut])
def update_fruit_purchase(
    purchase_id: int,
    req: FruitPurchaseUpdate,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    purchase = db.query(FruitPurchase).filter(FruitPurchase.id == purchase_id).first()
    if not purchase:
        raise HTTPException(status_code=404, detail="采购记录不存在")
    for k, v in req.model_dump(exclude_unset=True).items():
        setattr(purchase, k, v)
    db.commit()
    db.refresh(purchase)
    return ApiResponse(data=purchase)


@router.delete("/fruit/{purchase_id}", response_model=ApiResponse)
def delete_fruit_purchase(
    purchase_id: int,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    purchase = db.query(FruitPurchase).filter(FruitPurchase.id == purchase_id, FruitPurchase.deleted_at.is_(None)).first()
    if not purchase:
        raise HTTPException(status_code=404, detail="采购记录不存在")
    from datetime import datetime
    purchase.deleted_at = datetime.now()
    log_action(db, user, f"删除水果采购 #{purchase_id}（移入回收站）")
    db.commit()
    return ApiResponse(message="已移入回收站")


# ─── 材料采购 ───
@router.get("/material", response_model=PaginatedResponse[MaterialPurchaseOut])
def list_material_purchases(
    page: int = 1,
    page_size: int = 20,
    supplier_name: str | None = None,
    material_type: str | None = None,
    material_name: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    payment_status: str | None = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(SimpleMaterialPurchase).filter(SimpleMaterialPurchase.deleted_at.is_(None))
    if supplier_name:
        q = q.filter(SimpleMaterialPurchase.supplier_name.like(f"%{supplier_name}%"))
    if material_type:
        q = q.filter(SimpleMaterialPurchase.material_type == material_type)
    if material_name:
        q = q.filter(SimpleMaterialPurchase.material_name.like(f"%{material_name}%"))
    if start_date:
        q = q.filter(SimpleMaterialPurchase.purchase_date >= start_date)
    if end_date:
        q = q.filter(SimpleMaterialPurchase.purchase_date <= end_date)
    if payment_status:
        q = q.filter(SimpleMaterialPurchase.payment_status == payment_status)

    total = q.count()
    items = q.order_by(desc(SimpleMaterialPurchase.id)).offset((page - 1) * page_size).limit(page_size).all()
    return PaginatedResponse(data=items, total=total, page=page, page_size=page_size)


@router.post("/material", response_model=ApiResponse[MaterialPurchaseOut])
def create_material_purchase(
    req: MaterialPurchaseCreate,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if req.purchase_amount is not None and req.purchase_amount <= 0:
        raise HTTPException(status_code=400, detail="采购金额必须大于0")
    if req.purchase_amount is not None and req.purchase_amount > 1000000:
        raise HTTPException(status_code=400, detail="采购金额异常（超过100万），请检查输入")
    supplier = db.query(Supplier).filter(Supplier.id == req.supplier_id, Supplier.deleted_at.is_(None)).first()
    if not supplier:
        raise HTTPException(status_code=400, detail="供应商不存在")

    dup = db.query(SimpleMaterialPurchase).filter(
        SimpleMaterialPurchase.supplier_id == req.supplier_id,
        SimpleMaterialPurchase.material_name == req.material_name,
        SimpleMaterialPurchase.purchase_date == req.purchase_date,
        SimpleMaterialPurchase.purchase_amount == req.purchase_amount,
        SimpleMaterialPurchase.deleted_at.is_(None),
    ).first()
    if dup:
        raise HTTPException(status_code=409, detail=f"检测到重复采购记录（ID #{dup.id}，同日同供应商同材料同金额），如需继续请修改后重新提交")

    purchase = SimpleMaterialPurchase(**req.model_dump())
    db.add(purchase)
    log_action(db, user, f"新建材料采购：{req.material_name or req.material_type or ''} ¥{req.purchase_amount or 0}")
    db.commit()
    db.refresh(purchase)
    cache_clear_prefix("dashboard")

    finance_ok = True
    if req.purchase_amount and float(req.purchase_amount) > 0:
        try:
            result = push_payment_sync(
                supplier_name=req.supplier_name or "",
                amount=float(req.purchase_amount),
                reason=f"材料采购：{req.material_name or req.material_type or ''} ¥{req.purchase_amount}",
                source_order_id=purchase.id,
                source_order_type="material",
            )
            finance_ok = result.get("success", False)
        except Exception as e:
            finance_logger.error(f"Push material payment to finance failed: {e}")
            finance_ok = False

    msg = "创建成功"
    if not finance_ok:
        msg += "（财务系统推送失败，请联系管理员）"
    return ApiResponse(data=purchase, message=msg)


@router.put("/material/{purchase_id}", response_model=ApiResponse[MaterialPurchaseOut])
def update_material_purchase(
    purchase_id: int,
    req: MaterialPurchaseUpdate,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    p = db.query(SimpleMaterialPurchase).filter(SimpleMaterialPurchase.id == purchase_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="记录不存在")
    for k, v in req.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    log_action(db, user, f"编辑材料采购 #{purchase_id}")
    db.commit()
    db.refresh(p)
    return ApiResponse(data=p)


@router.delete("/material/{purchase_id}", response_model=ApiResponse)
def delete_material_purchase(
    purchase_id: int,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    p = db.query(SimpleMaterialPurchase).filter(SimpleMaterialPurchase.id == purchase_id, SimpleMaterialPurchase.deleted_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="记录不存在")
    from datetime import datetime
    p.deleted_at = datetime.now()
    log_action(db, user, f"删除材料采购 #{purchase_id}（移入回收站）")
    db.commit()
    return ApiResponse(message="已移入回收站")


# ─── 纸箱采购 ───
@router.get("/carton", response_model=PaginatedResponse[CartonPurchaseOut])
def list_carton_purchases(
    page: int = 1,
    page_size: int = 20,
    payment_status: str | None = None,
    supplier_id: int | None = None,
    carton_box_id: int | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = (
        db.query(
            CartonBoxPurchase,
            Supplier.name.label("supplier_name"),
            CartonBox.box_type.label("box_type"),
            CartonBox.stock_quantity.label("stock_quantity"),
        )
        .outerjoin(Supplier, CartonBoxPurchase.supplier_id == Supplier.id)
        .outerjoin(CartonBox, CartonBoxPurchase.carton_box_id == CartonBox.id)
        .filter(CartonBoxPurchase.deleted_at.is_(None))
    )
    if payment_status:
        q = q.filter(CartonBoxPurchase.payment_status == payment_status)
    if supplier_id:
        q = q.filter(CartonBoxPurchase.supplier_id == supplier_id)
    if carton_box_id:
        q = q.filter(CartonBoxPurchase.carton_box_id == carton_box_id)
    if start_date:
        q = q.filter(func.date(CartonBoxPurchase.created_at) >= start_date)
    if end_date:
        q = q.filter(func.date(CartonBoxPurchase.created_at) <= end_date)

    total = q.count()
    rows = q.order_by(desc(CartonBoxPurchase.id)).offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for purchase, sup_name, btype, stock_qty in rows:
        d = {c.name: getattr(purchase, c.name) for c in purchase.__table__.columns}
        d["supplier_name"] = sup_name
        d["box_type"] = btype
        d["stock_quantity"] = stock_qty
        items.append(d)

    return PaginatedResponse(data=items, total=total, page=page, page_size=page_size)


@router.get("/carton/stats")
def carton_purchase_stats(
    start_date: date | None = None,
    end_date: date | None = None,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    q = db.query(CartonBoxPurchase).filter(CartonBoxPurchase.deleted_at.is_(None))
    if start_date:
        q = q.filter(func.date(CartonBoxPurchase.created_at) >= start_date)
    if end_date:
        q = q.filter(func.date(CartonBoxPurchase.created_at) <= end_date)

    total_records = q.count()
    agg = q.with_entities(
        func.sum(CartonBoxPurchase.purchase_price * CartonBoxPurchase.purchase_quantity).label("total_amount"),
        func.sum(CartonBoxPurchase.purchase_quantity).label("total_qty"),
        func.count(func.distinct(CartonBoxPurchase.supplier_id)).label("supplier_count"),
        func.count(func.distinct(CartonBoxPurchase.carton_box_id)).label("box_type_count"),
    ).first()

    unpaid_agg = q.filter(CartonBoxPurchase.payment_status == "unpaid").with_entities(
        func.sum(CartonBoxPurchase.purchase_price * CartonBoxPurchase.purchase_quantity).label("unpaid_amount"),
        func.count().label("unpaid_count"),
    ).first()

    by_box = (
        q.with_entities(
            CartonBox.box_type,
            func.sum(CartonBoxPurchase.purchase_quantity).label("qty"),
            func.sum(CartonBoxPurchase.purchase_price * CartonBoxPurchase.purchase_quantity).label("amt"),
            func.avg(CartonBoxPurchase.purchase_price).label("avg_price"),
            CartonBox.stock_quantity,
            CartonBox.low_stock_threshold,
        )
        .outerjoin(CartonBox, CartonBoxPurchase.carton_box_id == CartonBox.id)
        .group_by(CartonBoxPurchase.carton_box_id, CartonBox.box_type, CartonBox.stock_quantity, CartonBox.low_stock_threshold)
        .all()
    )

    by_supplier = (
        q.with_entities(
            Supplier.name.label("supplier_name"),
            func.sum(CartonBoxPurchase.purchase_quantity).label("qty"),
            func.sum(CartonBoxPurchase.purchase_price * CartonBoxPurchase.purchase_quantity).label("amt"),
            func.count().label("order_count"),
        )
        .outerjoin(Supplier, CartonBoxPurchase.supplier_id == Supplier.id)
        .group_by(CartonBoxPurchase.supplier_id, Supplier.name)
        .order_by(desc("amt"))
        .all()
    )

    boxes = db.query(CartonBox).all()
    stock_overview = [
        {
            "id": b.id, "box_type": b.box_type,
            "stock_quantity": b.stock_quantity or 0,
            "low_stock_threshold": b.low_stock_threshold or 50,
            "is_low": (b.stock_quantity or 0) < (b.low_stock_threshold or 50),
            "purchase_price": float(b.purchase_price or 0),
        }
        for b in boxes
    ]

    return {
        "total_records": total_records,
        "total_amount": float(agg.total_amount or 0),
        "total_qty": int(agg.total_qty or 0),
        "supplier_count": int(agg.supplier_count or 0),
        "box_type_count": int(agg.box_type_count or 0),
        "unpaid_amount": float(unpaid_agg.unpaid_amount or 0) if unpaid_agg else 0,
        "unpaid_count": int(unpaid_agg.unpaid_count or 0) if unpaid_agg else 0,
        "by_box": [
            {
                "box_type": r.box_type or "未知",
                "qty": int(r.qty or 0),
                "amount": float(r.amt or 0),
                "avg_price": round(float(r.avg_price or 0), 2),
                "stock_quantity": int(r.stock_quantity or 0),
                "low_stock_threshold": int(r.low_stock_threshold or 50),
            }
            for r in by_box
        ],
        "by_supplier": [
            {"supplier_name": r.supplier_name or "未知", "qty": int(r.qty or 0), "amount": float(r.amt or 0), "order_count": int(r.order_count or 0)}
            for r in by_supplier
        ],
        "stock_overview": stock_overview,
    }


@router.get("/carton/price-trend")
def carton_price_trend(
    days: int = 90,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from datetime import timedelta
    cutoff = datetime.now() - timedelta(days=days)
    rows = (
        db.query(
            func.date(CartonBoxPurchase.created_at).label("purchase_date"),
            CartonBox.box_type,
            func.avg(CartonBoxPurchase.purchase_price).label("avg_price"),
            func.sum(CartonBoxPurchase.purchase_quantity).label("total_qty"),
        )
        .outerjoin(CartonBox, CartonBoxPurchase.carton_box_id == CartonBox.id)
        .filter(
            CartonBoxPurchase.deleted_at.is_(None),
            CartonBoxPurchase.created_at >= cutoff,
        )
        .group_by(func.date(CartonBoxPurchase.created_at), CartonBox.box_type)
        .order_by(func.date(CartonBoxPurchase.created_at))
        .all()
    )

    trend_data: dict = {}
    for r in rows:
        bt = r.box_type or "未知"
        if bt not in trend_data:
            trend_data[bt] = []
        trend_data[bt].append({
            "date": str(r.purchase_date),
            "avg_price": round(float(r.avg_price or 0), 2),
            "qty": int(r.total_qty or 0),
        })

    return {"days": days, "trends": trend_data}


@router.post("/carton/import")
async def import_carton_purchases(
    file: UploadFile = File(...),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="请选择文件")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail="仅支持 CSV 或 Excel 文件")

    content = await file.read()
    rows_data = []
    errors = []

    if ext == "csv":
        for encoding in ("utf-8-sig", "gbk", "gb2312", "utf-8"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise HTTPException(status_code=400, detail="文件编码无法识别")
        reader = csv.DictReader(io.StringIO(text))
        for i, row in enumerate(reader, 2):
            rows_data.append((i, row))
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                row_dict = {headers[j]: (str(v).strip() if v is not None else "") for j, v in enumerate(row) if j < len(headers)}
                rows_data.append((i, row_dict))
            wb.close()
        except ImportError:
            raise HTTPException(status_code=400, detail="请使用 CSV 格式")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"解析错误: {str(e)}")

    FIELD_MAP = {
        "供应商": "supplier_name", "supplier_name": "supplier_name", "供应商名称": "supplier_name",
        "纸箱规格": "box_type", "box_type": "box_type", "纸箱品种": "box_type", "规格": "box_type",
        "单价": "purchase_price", "purchase_price": "purchase_price", "采购单价": "purchase_price",
        "数量": "purchase_quantity", "purchase_quantity": "purchase_quantity", "采购数量": "purchase_quantity",
        "付款状态": "payment_status", "payment_status": "payment_status",
    }

    supplier_cache = {s.name: s.id for s in db.query(Supplier).filter(Supplier.deleted_at.is_(None), Supplier.type == "box").all()}
    box_cache = {b.box_type: b for b in db.query(CartonBox).all()}

    created = 0
    for line_no, raw_row in rows_data:
        row = {}
        for k, v in raw_row.items():
            mapped = FIELD_MAP.get(k.strip())
            if mapped:
                row[mapped] = v.strip() if isinstance(v, str) else v

        sup_name = row.get("supplier_name", "")
        box_type = row.get("box_type", "")
        if not sup_name:
            errors.append(f"第{line_no}行: 缺少供应商名称")
            continue
        if not box_type:
            errors.append(f"第{line_no}行: 缺少纸箱规格")
            continue

        sup_id = supplier_cache.get(sup_name)
        if not sup_id:
            errors.append(f"第{line_no}行: 供应商 '{sup_name}' 不存在")
            continue
        box = box_cache.get(box_type)
        if not box:
            errors.append(f"第{line_no}行: 纸箱规格 '{box_type}' 不存在")
            continue

        try:
            price = Decimal(str(row.get("purchase_price", "0")))
            qty = int(str(row.get("purchase_quantity", "0")))
            if price <= 0 or qty <= 0:
                errors.append(f"第{line_no}行: 单价或数量必须大于0")
                continue
        except (InvalidOperation, ValueError):
            errors.append(f"第{line_no}行: 单价或数量格式不正确")
            continue

        payment_status = "unpaid"
        ps_raw = str(row.get("payment_status", "")).lower().strip()
        if ps_raw in ("paid", "已付", "已付款"):
            payment_status = "paid"

        purchase = CartonBoxPurchase(
            supplier_id=sup_id,
            carton_box_id=box.id,
            purchase_price=price,
            purchase_quantity=qty,
            payment_status=payment_status,
        )
        db.add(purchase)
        box.stock_quantity = (box.stock_quantity or 0) + qty
        created += 1

    if created > 0:
        log_action(db, user, f"批量导入纸箱采购 {created} 条（文件: {file.filename}）")
        db.commit()
        cache_clear_prefix("dashboard")

    return ApiResponse(data={
        "created": created,
        "errors": errors[:20],
        "total_rows": len(rows_data),
        "error_count": len(errors),
    })


@router.get("/carton/import-template")
def carton_import_template(user: User = Depends(require_admin)):
    return ApiResponse(data={
        "headers": ["供应商名称", "纸箱规格", "采购单价", "采购数量", "付款状态"],
        "example": ["某纸箱厂", "5斤箱", "3.50", "1000", "unpaid"],
        "notes": "供应商名称和纸箱规格必须与系统中已有的一致；付款状态填 paid 或 unpaid",
    })


@router.post("/carton", response_model=ApiResponse[CartonPurchaseOut])
def create_carton_purchase(
    req: CartonPurchaseCreate,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if req.purchase_price <= 0:
        raise HTTPException(status_code=400, detail="采购单价必须大于0")
    if req.purchase_quantity <= 0:
        raise HTTPException(status_code=400, detail="采购数量必须大于0")
    if req.purchase_quantity > 100000:
        raise HTTPException(status_code=400, detail="采购数量异常（超过10万），请检查输入")
    box = db.query(CartonBox).filter(CartonBox.id == req.carton_box_id).first()
    if not box:
        raise HTTPException(status_code=400, detail="纸箱规格不存在")
    supplier = db.query(Supplier).filter(Supplier.id == req.supplier_id, Supplier.deleted_at.is_(None)).first()
    if not supplier:
        raise HTTPException(status_code=400, detail="供应商不存在")

    from datetime import datetime, timedelta
    five_min_ago = datetime.now() - timedelta(minutes=5)
    dup = db.query(CartonBoxPurchase).filter(
        CartonBoxPurchase.supplier_id == req.supplier_id,
        CartonBoxPurchase.carton_box_id == req.carton_box_id,
        CartonBoxPurchase.purchase_quantity == req.purchase_quantity,
        CartonBoxPurchase.purchase_price == req.purchase_price,
        CartonBoxPurchase.created_at >= five_min_ago,
        CartonBoxPurchase.deleted_at.is_(None),
    ).first()
    if dup:
        raise HTTPException(status_code=409, detail=f"检测到重复提交（5分钟内相同采购记录 #{dup.id}），请勿重复操作")

    purchase = CartonBoxPurchase(**req.model_dump())
    db.add(purchase)
    box.stock_quantity = (box.stock_quantity or 0) + req.purchase_quantity
    log_action(db, user, f"新建纸箱采购：{req.purchase_quantity}个 ¥{req.purchase_price}/个")
    db.commit()
    db.refresh(purchase)
    cache_clear_prefix("dashboard")

    total_amount = float(req.purchase_price * req.purchase_quantity)
    finance_ok = False
    try:
        result = push_payment_sync(
            supplier_name=supplier.name,
            amount=total_amount,
            reason=f"纸箱采购：{box.box_type} {req.purchase_quantity}个 ¥{req.purchase_price}/个",
            source_order_id=purchase.id,
            source_order_type="carton",
        )
        finance_ok = result.get("success", False)
    except Exception as e:
        finance_logger.error(f"Push carton payment to finance failed: {e}")

    msg = "创建成功"
    if not finance_ok:
        msg += "（财务系统推送失败，请联系管理员）"
    return ApiResponse(data=purchase, message=msg)


@router.put("/carton/{purchase_id}", response_model=ApiResponse[CartonPurchaseOut])
def update_carton_purchase(
    purchase_id: int,
    req: CartonPurchaseUpdate,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    p = db.query(CartonBoxPurchase).filter(CartonBoxPurchase.id == purchase_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="采购记录不存在")
    old_qty = p.purchase_quantity
    old_box_id = p.carton_box_id
    for k, v in req.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    if "purchase_quantity" in req.model_dump(exclude_unset=True) or "carton_box_id" in req.model_dump(exclude_unset=True):
        if old_box_id == p.carton_box_id:
            diff = p.purchase_quantity - old_qty
            box = db.query(CartonBox).filter(CartonBox.id == p.carton_box_id).first()
            if box:
                box.stock_quantity = max(0, (box.stock_quantity or 0) + diff)
        else:
            old_box = db.query(CartonBox).filter(CartonBox.id == old_box_id).first()
            if old_box:
                old_box.stock_quantity = max(0, (old_box.stock_quantity or 0) - old_qty)
            new_box = db.query(CartonBox).filter(CartonBox.id == p.carton_box_id).first()
            if new_box:
                new_box.stock_quantity = (new_box.stock_quantity or 0) + p.purchase_quantity
    log_action(db, user, f"编辑纸箱采购 #{purchase_id}")
    db.commit()
    db.refresh(p)
    return ApiResponse(data=p)


@router.delete("/carton/{purchase_id}", response_model=ApiResponse)
def delete_carton_purchase(
    purchase_id: int,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    p = db.query(CartonBoxPurchase).filter(CartonBoxPurchase.id == purchase_id, CartonBoxPurchase.deleted_at.is_(None)).first()
    if not p:
        raise HTTPException(status_code=404, detail="采购记录不存在")
    box = db.query(CartonBox).filter(CartonBox.id == p.carton_box_id).first()
    if box:
        box.stock_quantity = max(0, (box.stock_quantity or 0) - p.purchase_quantity)
    from datetime import datetime
    p.deleted_at = datetime.now()
    log_action(db, user, f"删除纸箱采购 #{purchase_id}（移入回收站）")
    db.commit()
    return ApiResponse(message="已移入回收站")


# ─── 批量更新付款状态（已弃用 — 付款状态由财务系统回调更新）───
@router.put("/payment-status", response_model=ApiResponse, deprecated=True)
def update_payment_status(
    req: PaymentStatusUpdate,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    raise HTTPException(
        status_code=403,
        detail="付款状态已改由财务系统管理，请通过财务系统操作付款",
    )


# ─── 批量导入水果采购 ───
@router.post("/fruit/import")
async def import_fruit_purchases(
    file: UploadFile = File(...),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="请选择文件")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail="仅支持 CSV 或 Excel 文件")

    content = await file.read()

    rows_data = []
    errors = []

    if ext == "csv":
        try:
            for encoding in ("utf-8-sig", "gbk", "gb2312", "utf-8"):
                try:
                    text = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise HTTPException(status_code=400, detail="文件编码无法识别，请使用 UTF-8 编码")

            reader = csv.DictReader(io.StringIO(text))
            for i, row in enumerate(reader, 2):
                rows_data.append((i, row))
        except csv.Error as e:
            raise HTTPException(status_code=400, detail=f"CSV 解析错误: {str(e)}")
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                row_dict = {headers[j]: (str(v).strip() if v is not None else "") for j, v in enumerate(row) if j < len(headers)}
                rows_data.append((i, row_dict))
            wb.close()
        except ImportError:
            raise HTTPException(status_code=400, detail="服务器未安装 openpyxl，请使用 CSV 格式")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Excel 解析错误: {str(e)}")

    FIELD_MAP = {
        "水果名称": "fruit_name", "fruit_name": "fruit_name",
        "供应商": "supplier_name", "supplier_name": "supplier_name", "供应商名称": "supplier_name",
        "采购日期": "purchase_date", "purchase_date": "purchase_date", "日期": "purchase_date",
        "单价": "purchase_price", "purchase_price": "purchase_price", "采购单价": "purchase_price", "价格": "purchase_price",
        "重量": "purchase_weight", "purchase_weight": "purchase_weight", "采购重量": "purchase_weight", "kg": "purchase_weight",
        "付款状态": "payment_status", "payment_status": "payment_status",
    }

    fruit_cache = {f.name: f.id for f in db.query(Fruit.name, Fruit.id).all()}
    supplier_cache = {}
    for s in db.query(Supplier.name, Supplier.id, Supplier.type).filter(Supplier.deleted_at.is_(None)).all():
        supplier_cache[s.name] = s.id

    created = 0
    for line_no, raw_row in rows_data:
        row = {}
        for k, v in raw_row.items():
            mapped = FIELD_MAP.get(k.strip())
            if mapped:
                row[mapped] = v.strip() if isinstance(v, str) else v

        if not row.get("fruit_name"):
            errors.append(f"第{line_no}行: 缺少水果名称")
            continue
        if not row.get("supplier_name"):
            errors.append(f"第{line_no}行: 缺少供应商名称")
            continue

        try:
            price = Decimal(str(row.get("purchase_price", "0")))
            weight = Decimal(str(row.get("purchase_weight", "0")))
            if price <= 0 or weight <= 0:
                errors.append(f"第{line_no}行: 单价或重量必须大于0")
                continue
        except (InvalidOperation, ValueError):
            errors.append(f"第{line_no}行: 单价或重量格式不正确")
            continue

        raw_date = row.get("purchase_date", "")
        purchase_date = None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                purchase_date = datetime.strptime(str(raw_date), fmt).date()
                break
            except ValueError:
                continue
        if not purchase_date:
            try:
                purchase_date = date.fromisoformat(str(raw_date))
            except (ValueError, TypeError):
                errors.append(f"第{line_no}行: 日期格式不正确 ({raw_date})")
                continue

        fruit_name = row["fruit_name"]
        supplier_name = row["supplier_name"]
        fruit_id = fruit_cache.get(fruit_name, 0)
        supplier_id = supplier_cache.get(supplier_name, 0)

        payment_status = "unpaid"
        ps_raw = str(row.get("payment_status", "")).lower().strip()
        if ps_raw in ("paid", "已付", "已付款"):
            payment_status = "paid"

        purchase = FruitPurchase(
            supplier_id=supplier_id,
            fruit_id=fruit_id,
            supplier_name=supplier_name,
            fruit_name=fruit_name,
            purchase_date=purchase_date,
            purchase_price=price,
            purchase_weight=weight,
            payment_status=payment_status,
        )
        db.add(purchase)
        created += 1

    if created > 0:
        log_action(db, user, f"批量导入水果采购 {created} 条（文件: {file.filename}）")
        db.commit()
        cache_clear_prefix("dashboard")

    return ApiResponse(data={
        "created": created,
        "errors": errors[:20],
        "total_rows": len(rows_data),
        "error_count": len(errors),
    })


@router.get("/fruit/import-template")
def download_import_template(user: User = Depends(require_admin)):
    return ApiResponse(data={
        "headers": ["水果名称", "供应商名称", "采购日期", "采购单价", "采购重量", "付款状态"],
        "example": ["芒果", "张三水果批发", "2025-03-01", "5.50", "100", "unpaid"],
        "notes": "付款状态填 paid 或 unpaid；日期格式 YYYY-MM-DD；重量单位 kg",
    })


# ─── 材料采购批量导入 ───
@router.post("/material/import")
async def import_material_purchases(
    file: UploadFile = File(...),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="请选择文件")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail="仅支持 CSV 或 Excel 文件")

    content = await file.read()
    rows_data = []
    errors = []

    if ext == "csv":
        for encoding in ("utf-8-sig", "gbk", "gb2312", "utf-8"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise HTTPException(status_code=400, detail="文件编码无法识别")
        reader = csv.DictReader(io.StringIO(text))
        for i, row in enumerate(reader, 2):
            rows_data.append((i, row))
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                row_dict = {headers[j]: (str(v).strip() if v is not None else "") for j, v in enumerate(row) if j < len(headers)}
                rows_data.append((i, row_dict))
            wb.close()
        except ImportError:
            raise HTTPException(status_code=400, detail="请使用 CSV 格式")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"解析错误: {str(e)}")

    FIELD_MAP = {
        "材料名称": "material_name", "material_name": "material_name", "名称": "material_name",
        "材料类型": "material_type", "material_type": "material_type", "类型": "material_type",
        "供应商": "supplier_name", "supplier_name": "supplier_name", "供应商名称": "supplier_name",
        "采购日期": "purchase_date", "purchase_date": "purchase_date", "日期": "purchase_date",
        "金额": "purchase_amount", "purchase_amount": "purchase_amount", "采购金额": "purchase_amount",
        "付款状态": "payment_status", "payment_status": "payment_status",
        "备注": "notes", "notes": "notes",
    }

    supplier_cache = {}
    for s in db.query(Supplier.name, Supplier.id).filter(Supplier.deleted_at.is_(None), Supplier.type == "material").all():
        supplier_cache[s.name] = s.id

    created = 0
    for line_no, raw_row in rows_data:
        row = {}
        for k, v in raw_row.items():
            mapped = FIELD_MAP.get(k.strip())
            if mapped:
                row[mapped] = v.strip() if isinstance(v, str) else v

        if not row.get("material_name") and not row.get("material_type"):
            errors.append(f"第{line_no}行: 缺少材料名称或类型")
            continue
        if not row.get("supplier_name"):
            errors.append(f"第{line_no}行: 缺少供应商名称")
            continue

        try:
            amount = Decimal(str(row.get("purchase_amount", "0")))
            if amount <= 0:
                errors.append(f"第{line_no}行: 金额必须大于0")
                continue
        except (InvalidOperation, ValueError):
            errors.append(f"第{line_no}行: 金额格式不正确")
            continue

        raw_date = row.get("purchase_date", "")
        purchase_date = None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                purchase_date = datetime.strptime(str(raw_date), fmt).date()
                break
            except ValueError:
                continue
        if not purchase_date:
            try:
                purchase_date = date.fromisoformat(str(raw_date))
            except (ValueError, TypeError):
                errors.append(f"第{line_no}行: 日期格式不正确 ({raw_date})")
                continue

        supplier_name = row["supplier_name"]
        supplier_id = supplier_cache.get(supplier_name, 0)

        payment_status = "unpaid"
        ps_raw = str(row.get("payment_status", "")).lower().strip()
        if ps_raw in ("paid", "已付", "已付款"):
            payment_status = "paid"

        purchase = SimpleMaterialPurchase(
            supplier_id=supplier_id,
            supplier_name=supplier_name,
            material_type=row.get("material_type", ""),
            material_name=row.get("material_name", ""),
            purchase_amount=amount,
            purchase_date=purchase_date,
            payment_status=payment_status,
            notes=row.get("notes", ""),
        )
        db.add(purchase)
        created += 1

    if created > 0:
        log_action(db, user, f"批量导入材料采购 {created} 条（文件: {file.filename}）")
        db.commit()
        cache_clear_prefix("dashboard")

    return ApiResponse(data={
        "created": created,
        "errors": errors[:20],
        "total_rows": len(rows_data),
        "error_count": len(errors),
    })


@router.get("/material/import-template")
def material_import_template(user: User = Depends(require_admin)):
    return ApiResponse(data={
        "headers": ["材料名称", "材料类型", "供应商名称", "采购日期", "采购金额", "付款状态", "备注"],
        "example": ["胶带", "包装材料", "某材料商", "2025-03-01", "500", "unpaid", ""],
        "notes": "付款状态填 paid 或 unpaid；日期格式 YYYY-MM-DD；金额为总金额",
    })


@router.get("/material/stats")
def material_purchase_stats(
    start_date: date | None = None,
    end_date: date | None = None,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    q = db.query(SimpleMaterialPurchase).filter(SimpleMaterialPurchase.deleted_at.is_(None))
    if start_date:
        q = q.filter(SimpleMaterialPurchase.purchase_date >= start_date)
    if end_date:
        q = q.filter(SimpleMaterialPurchase.purchase_date <= end_date)

    total_records = q.count()
    agg = q.with_entities(
        func.sum(SimpleMaterialPurchase.purchase_amount).label("total_amount"),
        func.count(func.distinct(SimpleMaterialPurchase.supplier_id)).label("supplier_count"),
    ).first()

    unpaid_agg = q.filter(SimpleMaterialPurchase.payment_status == "unpaid").with_entities(
        func.sum(SimpleMaterialPurchase.purchase_amount).label("unpaid_amount"),
        func.count().label("unpaid_count"),
    ).first()

    by_type = (
        q.with_entities(
            SimpleMaterialPurchase.material_type,
            func.count().label("cnt"),
            func.sum(SimpleMaterialPurchase.purchase_amount).label("amt"),
        )
        .group_by(SimpleMaterialPurchase.material_type)
        .order_by(desc("amt"))
        .all()
    )

    by_supplier = (
        q.with_entities(
            SimpleMaterialPurchase.supplier_name,
            func.count().label("cnt"),
            func.sum(SimpleMaterialPurchase.purchase_amount).label("amt"),
        )
        .group_by(SimpleMaterialPurchase.supplier_name)
        .order_by(desc("amt"))
        .all()
    )

    monthly = (
        q.with_entities(
            func.date_format(SimpleMaterialPurchase.purchase_date, "%Y-%m").label("month"),
            func.sum(SimpleMaterialPurchase.purchase_amount).label("amt"),
            func.count().label("cnt"),
        )
        .filter(SimpleMaterialPurchase.purchase_date.isnot(None))
        .group_by("month")
        .order_by("month")
        .all()
    )

    return {
        "total_records": total_records,
        "total_amount": float(agg.total_amount or 0) if agg else 0,
        "supplier_count": int(agg.supplier_count or 0) if agg else 0,
        "unpaid_amount": float(unpaid_agg.unpaid_amount or 0) if unpaid_agg else 0,
        "unpaid_count": int(unpaid_agg.unpaid_count or 0) if unpaid_agg else 0,
        "by_type": [
            {"type": r.material_type or "未分类", "count": int(r.cnt or 0), "amount": float(r.amt or 0)}
            for r in by_type
        ],
        "by_supplier": [
            {"supplier_name": r.supplier_name or "未知", "count": int(r.cnt or 0), "amount": float(r.amt or 0)}
            for r in by_supplier
        ],
        "monthly": [
            {"month": r.month, "amount": float(r.amt or 0), "count": int(r.cnt or 0)}
            for r in monthly
        ],
    }


# ─── 批量删除 ───
from pydantic import BaseModel
from typing import List

class BatchDeleteRequest(BaseModel):
    order_type: str
    order_ids: List[int]


@router.post("/batch-delete", response_model=ApiResponse)
def batch_delete_orders(
    req: BatchDeleteRequest,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if not req.order_ids:
        raise HTTPException(status_code=400, detail="请选择要删除的订单")
    if len(req.order_ids) > 100:
        raise HTTPException(status_code=400, detail="单次最多删除100条")

    now = datetime.now()
    count = 0

    if req.order_type == "fruit":
        records = db.query(FruitPurchase).filter(
            FruitPurchase.id.in_(req.order_ids),
            FruitPurchase.deleted_at.is_(None),
        ).all()
        for r in records:
            r.deleted_at = now
            count += 1
    elif req.order_type == "carton":
        records = db.query(CartonBoxPurchase).filter(
            CartonBoxPurchase.id.in_(req.order_ids),
            CartonBoxPurchase.deleted_at.is_(None),
        ).all()
        for r in records:
            r.deleted_at = now
            count += 1
    elif req.order_type == "material":
        records = db.query(SimpleMaterialPurchase).filter(
            SimpleMaterialPurchase.id.in_(req.order_ids),
            SimpleMaterialPurchase.deleted_at.is_(None),
        ).all()
        for r in records:
            r.deleted_at = now
            count += 1
    else:
        raise HTTPException(status_code=400, detail="无效的订单类型")

    if count > 0:
        type_name = {"fruit": "水果", "carton": "纸箱", "material": "材料"}.get(req.order_type, req.order_type)
        log_action(db, user, f"批量删除{type_name}采购 {count} 条（移入回收站）")
        db.commit()
        cache_clear_prefix("dashboard")

    return ApiResponse(message=f"已将 {count} 条订单移入回收站", data={"deleted": count})


@router.post("/migrate-all-paid", response_model=ApiResponse)
def migrate_all_unpaid_to_paid(
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """一键将所有历史待付订单标记为已付（不推送财务系统）"""
    fruit_count = db.query(FruitPurchase).filter(
        FruitPurchase.payment_status == "unpaid",
        FruitPurchase.deleted_at.is_(None),
    ).update({"payment_status": "paid"}, synchronize_session="fetch")

    material_count = db.query(SimpleMaterialPurchase).filter(
        SimpleMaterialPurchase.payment_status == "unpaid",
        SimpleMaterialPurchase.deleted_at.is_(None),
    ).update({"payment_status": "paid"}, synchronize_session="fetch")

    carton_count = db.query(CartonBoxPurchase).filter(
        CartonBoxPurchase.payment_status == "unpaid",
        CartonBoxPurchase.deleted_at.is_(None),
    ).update({"payment_status": "paid"}, synchronize_session="fetch")

    total = fruit_count + material_count + carton_count
    log_action(db, user, f"一键标记历史已付：水果{fruit_count}条、材料{material_count}条、纸箱{carton_count}条")
    db.commit()
    cache_clear_prefix("dashboard")
    return ApiResponse(
        message=f"已将 {total} 条历史订单标记为已付",
        data={"fruit": fruit_count, "material": material_count, "carton": carton_count, "total": total},
    )


@router.put("/payment-callback", response_model=ApiResponse)
def payment_callback(
    source_order_type: str = Query(...),
    source_order_id: int = Query(...),
    status: str = Query(...),
    trade_no: str = Query(None),
    api_key: str = Query(...),
    db: Session = Depends(get_db),
):
    """财务系统付款完成后的回调接口"""
    if api_key != "fruit-admin-bridge-2026":
        raise HTTPException(status_code=403, detail="Invalid API key")
    if status not in ("paid", "unpaid"):
        raise HTTPException(status_code=400, detail="Invalid status")

    model_map = {
        "fruit": FruitPurchase,
        "material": SimpleMaterialPurchase,
        "carton": CartonBoxPurchase,
    }
    model = model_map.get(source_order_type)
    if not model:
        raise HTTPException(status_code=400, detail="Invalid order type")

    record = db.query(model).filter(model.id == source_order_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Order not found")

    record.payment_status = status
    db.commit()
    return ApiResponse(message="Payment status updated")


@router.get("/fruit/batch-progress")
def fruit_batch_progress(
    page: int = 1,
    page_size: int = 20,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """水果采购批次进度追踪"""
    from app.models import BatchAssignment, SkuTransaction, PrintedLabel, WorkerProduction
    from sqlalchemy import Integer, case

    q = db.query(FruitPurchase).filter(FruitPurchase.deleted_at.is_(None))
    total = q.count()
    purchases = q.order_by(desc(FruitPurchase.id)).offset((page - 1) * page_size).limit(page_size).all()

    if not purchases:
        return ApiResponse(data={"items": [], "total": total, "page": page, "page_size": page_size})

    ids = [p.id for p in purchases]

    assign_stats = {}
    for r in db.query(
        BatchAssignment.purchase_id,
        func.count(func.distinct(BatchAssignment.worker_id)).label("workers"),
    ).filter(BatchAssignment.purchase_id.in_(ids)).group_by(BatchAssignment.purchase_id).all():
        assign_stats[r.purchase_id] = int(r.workers)

    txn_stats = {}
    for r in db.query(
        SkuTransaction.fruit_purchase_id,
        func.count(SkuTransaction.id).label("txn_count"),
        func.coalesce(func.sum(SkuTransaction.quantity), 0).label("txn_qty"),
        func.sum(case((SkuTransaction.is_printed == True, 1), else_=0)).label("printed_txns"),
    ).filter(SkuTransaction.fruit_purchase_id.in_(ids)).group_by(SkuTransaction.fruit_purchase_id).all():
        txn_stats[r.fruit_purchase_id] = {
            "txn_count": int(r.txn_count), "txn_qty": int(r.txn_qty),
            "printed_txns": int(r.printed_txns or 0),
        }

    label_stats = {}
    for r in db.query(
        PrintedLabel.b,
        func.count(PrintedLabel.id).label("total"),
        func.sum(func.cast(PrintedLabel.scanned_outbound > 0, Integer)).label("outbound"),
    ).filter(PrintedLabel.b.in_(ids)).group_by(PrintedLabel.b).all():
        label_stats[r.b] = {"total": int(r.total), "outbound": int(r.outbound or 0)}

    prod_stats = {}
    for r in db.query(
        WorkerProduction.sku_id,
        func.sum(WorkerProduction.actual_packaging_quantity).label("total_qty"),
        func.sum(case((WorkerProduction.audit_status == "approved", WorkerProduction.actual_packaging_quantity), else_=0)).label("approved_qty"),
    ).join(SkuTransaction, (SkuTransaction.sku_id == WorkerProduction.sku_id)).filter(
        SkuTransaction.fruit_purchase_id.in_(ids)
    ).group_by(WorkerProduction.sku_id).all():
        pass

    items = []
    for p in purchases:
        pid = p.id
        assigns = assign_stats.get(pid, 0)
        txns = txn_stats.get(pid, {"txn_count": 0, "txn_qty": 0, "printed_txns": 0})
        labels = label_stats.get(pid, {"total": 0, "outbound": 0})

        total_labels = labels["total"]
        outbound_labels = labels["outbound"]
        instock = total_labels - outbound_labels

        if total_labels == 0 and assigns == 0:
            stage = "new"
        elif assigns > 0 and total_labels == 0:
            stage = "assigned"
        elif total_labels > 0 and outbound_labels == 0:
            stage = "producing"
        elif outbound_labels > 0 and outbound_labels < total_labels:
            stage = "shipping"
        elif total_labels > 0 and outbound_labels >= total_labels:
            stage = "completed"
        else:
            stage = "new"

        outbound_rate = round(outbound_labels / max(total_labels, 1) * 100, 1)

        items.append({
            "id": pid,
            "fruit_name": p.fruit_name,
            "supplier_name": p.supplier_name,
            "purchase_date": p.purchase_date.isoformat() if p.purchase_date else None,
            "purchase_weight": float(p.purchase_weight or 0),
            "total_amount": round(float(p.purchase_price or 0) * float(p.purchase_weight or 0), 2),
            "payment_status": p.payment_status,
            "assigned_workers": assigns,
            "sku_requests": txns["txn_count"],
            "request_qty": txns["txn_qty"],
            "total_labels": total_labels,
            "outbound_labels": outbound_labels,
            "instock_labels": instock,
            "outbound_rate": outbound_rate,
            "stage": stage,
        })

    return ApiResponse(data={"items": items, "total": total, "page": page, "page_size": page_size})


@router.get("/purchase-intelligence")
def purchase_intelligence(
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """采购智能分析 — 基于历史数据的采购建议"""
    from app.models import PrintedLabel, Sku
    from sqlalchemy import case
    today = date.today()
    d30 = today - timedelta(days=30)
    d60 = today - timedelta(days=60)

    def _f(v):
        return float(v) if isinstance(v, Decimal) else (v or 0)

    fruit_stats = db.query(
        Fruit.id.label("fruit_id"), Fruit.name.label("fruit_name"),
        func.count(FruitPurchase.id).label("purchase_count"),
        func.coalesce(func.sum(FruitPurchase.purchase_weight), 0).label("total_weight"),
        func.coalesce(func.sum(FruitPurchase.purchase_price * FruitPurchase.purchase_weight), 0).label("total_cost"),
        func.coalesce(func.avg(FruitPurchase.purchase_price), 0).label("avg_price"),
        func.max(FruitPurchase.purchase_date).label("last_purchase"),
    ).outerjoin(FruitPurchase, (FruitPurchase.fruit_id == Fruit.id) & (FruitPurchase.deleted_at.is_(None)) & (FruitPurchase.purchase_date >= d60)).group_by(Fruit.id, Fruit.name).all()

    recent_prices = {}
    for fp in db.query(FruitPurchase.fruit_name, FruitPurchase.purchase_price, FruitPurchase.purchase_date).filter(
        FruitPurchase.deleted_at.is_(None), FruitPurchase.purchase_date >= d30).order_by(FruitPurchase.purchase_date.desc()).all():
        fn = fp.fruit_name or "未知"
        if fn not in recent_prices:
            recent_prices[fn] = []
        if len(recent_prices[fn]) < 5:
            recent_prices[fn].append({"price": round(_f(fp.purchase_price), 2), "date": str(fp.purchase_date)})

    label_consumption = {}
    for r in db.query(
        Sku.fruit_name,
        func.count(PrintedLabel.id).label("labels"),
        func.sum(case((PrintedLabel.scanned_outbound > 0, 1), else_=0)).label("outbound"),
    ).join(Sku, PrintedLabel.s == Sku.id).filter(
        func.date(PrintedLabel.created_at) >= d30).group_by(Sku.fruit_name).all():
        fn = r.fruit_name or "未知"
        label_consumption[fn] = {"labels": int(r.labels or 0), "outbound": int(r.outbound or 0)}

    supplier_stats = {}
    for r in db.query(
        FruitPurchase.supplier_name,
        func.count(FruitPurchase.id).label("count"),
        func.avg(FruitPurchase.purchase_price).label("avg_price"),
        func.sum(FruitPurchase.purchase_weight).label("total_weight"),
    ).filter(FruitPurchase.deleted_at.is_(None), FruitPurchase.purchase_date >= d60).group_by(
        FruitPurchase.supplier_name).all():
        sn = r.supplier_name or "未知"
        supplier_stats[sn] = {
            "count": int(r.count or 0),
            "avg_price": round(_f(r.avg_price), 2),
            "total_weight": round(_f(r.total_weight), 1),
        }

    fruits_data = []
    for fs in fruit_stats:
        fn = fs.fruit_name
        consumption = label_consumption.get(fn, {"labels": 0, "outbound": 0})
        days_since = (today - fs.last_purchase).days if fs.last_purchase else 999
        daily_consumption = consumption["outbound"] / 30 if consumption["outbound"] > 0 else 0

        urgency = "low"
        if days_since > 14 and daily_consumption > 5:
            urgency = "high"
        elif days_since > 7 and daily_consumption > 2:
            urgency = "medium"
        elif days_since > 21:
            urgency = "medium"

        fruits_data.append({
            "fruit_id": fs.fruit_id, "fruit_name": fn,
            "purchase_count": int(fs.purchase_count or 0),
            "total_weight": round(_f(fs.total_weight), 1),
            "total_cost": round(_f(fs.total_cost), 0),
            "avg_price": round(_f(fs.avg_price), 2),
            "last_purchase": str(fs.last_purchase) if fs.last_purchase else None,
            "days_since_last": days_since,
            "monthly_labels": consumption["labels"],
            "monthly_outbound": consumption["outbound"],
            "daily_consumption": round(daily_consumption, 1),
            "recent_prices": recent_prices.get(fn, []),
            "urgency": urgency,
        })

    fruits_data.sort(key=lambda x: {"high": 0, "medium": 1, "low": 2}[x["urgency"]])

    suppliers = sorted(supplier_stats.items(), key=lambda x: x[1]["count"], reverse=True)
    supplier_list = [{"name": k, **v} for k, v in suppliers[:10]]

    return ApiResponse(data={
        "fruits": fruits_data,
        "suppliers": supplier_list,
        "summary": {
            "total_fruits": len(fruits_data),
            "high_urgency": sum(1 for f in fruits_data if f["urgency"] == "high"),
            "medium_urgency": sum(1 for f in fruits_data if f["urgency"] == "medium"),
            "total_suppliers": len(supplier_list),
        },
    })


@router.get("/purchase-ai-suggest")
def purchase_ai_suggest(
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """AI采购建议 — 流式生成"""
    from fastapi.responses import StreamingResponse
    from app.models import PrintedLabel, Sku
    from sqlalchemy import case
    import json
    today = date.today()
    d30 = today - timedelta(days=30)

    def _f(v):
        return float(v) if isinstance(v, Decimal) else (v or 0)

    ctx = [f"日期: {today}"]

    fruit_q = db.query(
        Fruit.name, func.count(FruitPurchase.id), func.avg(FruitPurchase.purchase_price),
        func.sum(FruitPurchase.purchase_weight), func.max(FruitPurchase.purchase_date),
    ).outerjoin(FruitPurchase, (FruitPurchase.fruit_id == Fruit.id) & (FruitPurchase.deleted_at.is_(None)) & (FruitPurchase.purchase_date >= d30)).group_by(Fruit.id, Fruit.name).all()

    ctx.append("水果采购情况(近30天):")
    for r in fruit_q:
        days_since = (today - r[4]).days if r[4] else 999
        ctx.append(f"  {r[0]}: {int(r[1] or 0)}笔, 均价¥{_f(r[2]):.2f}/kg, 共{_f(r[3]):.0f}kg, 距上次{days_since}天")

    consumption = db.query(Sku.fruit_name, func.count(PrintedLabel.id),
        func.sum(case((PrintedLabel.scanned_outbound > 0, 1), else_=0))).join(
        Sku, PrintedLabel.s == Sku.id).filter(func.date(PrintedLabel.created_at) >= d30).group_by(Sku.fruit_name).all()
    ctx.append("水果消耗(近30天):")
    for r in consumption:
        ctx.append(f"  {r[0] or '未知'}: 打印{int(r[1] or 0)}, 出库{int(r[2] or 0)}")

    prompt = f"""请根据以下采购和消耗数据，生成采购建议。

{chr(10).join(ctx)}

请用markdown格式回复，包含：
1. **采购优先级**: 哪些水果需要优先采购，为什么
2. **建议采购量**: 基于消耗速度推荐采购量
3. **价格参考**: 基于历史均价给出参考价格区间
4. **供应商建议**: 采购策略建议
5. **风险提示**: 需要注意的问题

简洁专业，不超过300字。"""

    def generate():
        try:
            from openai import OpenAI
            client = OpenAI(api_key="sk-b121d7a1020f4c4e9740ec130f359333",
                          base_url="https://dashscope.aliyuncs.com/compatible-mode/v1")
            resp = client.chat.completions.create(model="qwen-plus", messages=[
                {"role": "system", "content": "你是果管系统的采购顾问。基于数据给出精准采购建议。使用简体中文。"},
                {"role": "user", "content": prompt}], stream=True, temperature=0.2, max_tokens=2000)
            for chunk in resp:
                if chunk.choices and chunk.choices[0].delta.content:
                    yield f"data: {json.dumps({'content': chunk.choices[0].delta.content}, ensure_ascii=False)}\n\n"
            yield "data: [DONE]\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'error': str(e)}, ensure_ascii=False)}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})
